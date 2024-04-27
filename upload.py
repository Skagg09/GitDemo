import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait


def update_xml(filepath, colnName, searchTerm, newvalue):
    book = openpyxl.load_workbook(filepath)
    sheet = book.active
    Dict = {}
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colnName:
            Dict["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value=newvalue
    book.save(filepath)




driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID,"downloadButton").click()
time.sleep(3)
fp = "C:\\Users\\surya\\Downloads\\download.xlsx"
ftName="Apple"
nPrice=900
update_xml(fp,"price",ftName,nPrice)
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(fp)

wait = WebDriverWait(driver, 5)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.invisibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
priceColumn = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
time.sleep(10)
actual_price = driver.find_element(By.XPATH,"//div[text()='"+ftName+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
print(actual_price)
assert int(actual_price) == nPrice